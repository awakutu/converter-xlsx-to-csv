"""
Interactive XLSX to CSV Converter

Features:
- Interactive file selection dialog for input XLSX file
- User input for output CSV filename
- Stable numeric formatting (no scientific notation)
- Proper date/datetime formatting
- Progress bar with tqdm
- Clean string normalization
"""

import csv
import logging
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except ImportError:
    print("Error: tkinter is required for file dialogs")
    sys.exit(1)

# Configuration
DATE_FORMAT_DATETIME = "%d/%m/%Y, %H:%M:%S"
DATE_FORMAT_DATE = "%d/%m/%Y"
ENCODING = "utf-8-sig"

# Regex untuk menghapus karakter zero-width
_ZERO_WIDTH_RE = re.compile(r"[\u200B\u200C\u200D\uFEFF]")
NBSP = "\u00a0"


def setup_logging():
    """Setup basic logging"""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )


def to_plain_string(cell) -> str:
    """
    Convert openpyxl cell to stable string representation:
    - Dates/datetimes formatted properly
    - Numbers without scientific notation
    - Booleans as TRUE/FALSE
    - None as empty string
    """
    v = cell.value
    if v is None:
        return ""

    # Handle dates and datetimes
    if getattr(cell, "is_date", False) and isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.strftime(DATE_FORMAT_DATETIME)
        return v.strftime(DATE_FORMAT_DATE)

    # Strings as-is
    if isinstance(v, str):
        return v

    # Booleans
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"

    # Numbers - avoid scientific notation
    if isinstance(v, (int, float)):
        try:
            d = Decimal(str(v))
            fixed = format(d, "f")
            if "." in fixed:
                fixed = fixed.rstrip("0").rstrip(".")
            return fixed
        except (InvalidOperation, ValueError):
            return str(v)

    # Fallback
    return str(v)


def trim_text(s: str) -> str:
    """
    Normalize and clean text:
    - Convert NBSP to regular space
    - Remove zero-width characters
    - Strip whitespace
    """
    if not s:
        return ""
    if not isinstance(s, str):
        s = str(s)

    s = s.replace(NBSP, " ")
    s = _ZERO_WIDTH_RE.sub("", s)
    return s.strip()


def select_input_file() -> Optional[str]:
    """Open file dialog to select input XLSX file"""
    root = tk.Tk()
    root.withdraw()  # Hide main window

    file_path = filedialog.askopenfilename(
        title="Pilih file Excel (.xlsx) untuk dikonversi",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
    )

    root.destroy()
    return file_path if file_path else None


def get_output_filename(input_path: str) -> str:
    """Get output CSV filename from user input"""
    input_name = Path(input_path).stem
    default_output = f"{input_name}.csv"

    print(f"\nFile input: {input_path}")
    print(f"Nama output default: {default_output}")

    while True:
        output_name = input(
            f"Masukkan nama file output CSV (tekan Enter untuk '{default_output}'): "
        ).strip()

        if not output_name:
            output_name = default_output

        if not output_name.lower().endswith(".csv"):
            output_name += ".csv"

        # Check if file exists
        if os.path.exists(output_name):
            overwrite = input(f"File '{output_name}' sudah ada. Timpa? (y/n): ").lower()
            if overwrite in ["y", "yes"]:
                break
        else:
            break

    return output_name


def convert_xlsx_to_csv(input_path: str, output_path: str) -> int:
    """Convert XLSX to CSV and return number of rows processed"""
    logging.info(f"Membuka file: {input_path}")

    try:
        wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
        ws = wb.active

        # Get total rows for progress bar
        total_rows = ws.max_row or 0

        logging.info(f"Mengkonversi sheet: {ws.title}")
        logging.info(f"Menulis ke: {output_path}")

        rows_written = 0

        with open(output_path, "w", newline="", encoding=ENCODING) as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)

            # Setup iterator with optional progress bar
            iterator = ws.iter_rows(values_only=False)
            if tqdm:
                iterator = tqdm(iterator, total=total_rows, desc="Converting")

            for row_cells in iterator:
                out_row = []
                for cell in row_cells:
                    raw = to_plain_string(cell)
                    cleaned = trim_text(raw)
                    out_row.append(cleaned)

                writer.writerow(out_row)
                rows_written += 1

        wb.close()
        logging.info(f"Konversi selesai. {rows_written} baris ditulis.")
        return rows_written

    except Exception as e:
        logging.error(f"Error saat konversi: {e}")
        raise


def main():
    """Main function"""
    setup_logging()

    print("=== XLSX to CSV Converter ===")
    print("Converter ini akan mengubah file Excel (.xlsx) menjadi CSV")
    print()

    # Select input file
    print("1. Pilih file Excel input...")
    input_path = select_input_file()

    if not input_path:
        print("Tidak ada file yang dipilih. Program berhenti.")
        return

    if not os.path.exists(input_path):
        print(f"Error: File '{input_path}' tidak ditemukan.")
        return

    # Get output filename
    print("\n2. Tentukan nama file output...")
    output_path = get_output_filename(input_path)

    # Confirm conversion
    print(f"\n3. Konfirmasi konversi:")
    print(f"   Input:  {input_path}")
    print(f"   Output: {output_path}")

    confirm = input("\nLanjutkan konversi? (y/n): ").lower()
    if confirm not in ["y", "yes"]:
        print("Konversi dibatalkan.")
        return

    # Perform conversion
    print("\n4. Memulai konversi...")
    try:
        rows_count = convert_xlsx_to_csv(input_path, output_path)
        print(f"\n✅ Konversi berhasil!")
        print(f"   File output: {output_path}")
        print(f"   Total baris: {rows_count}")

        # Ask if user wants to convert another file
        print()
        another = input("Konversi file lain? (y/n): ").lower()
        if another in ["y", "yes"]:
            print("\n" + "=" * 50)
            main()  # Recursive call for another conversion

    except Exception as e:
        print(f"\n❌ Konversi gagal: {e}")
        logging.exception("Conversion failed")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProgram dihentikan oleh user.")
    except Exception as e:
        print(f"\nError tak terduga: {e}")
        logging.exception("Unexpected error")
    finally:
        input("\nTekan Enter untuk keluar...")
